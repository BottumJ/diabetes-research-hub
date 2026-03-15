"""
PROJECT 1: Literature Gap Analysis
Diabetes Research Hub — Research Doctrine Compliant

Queries PubMed E-utilities API to map publication density across all
research domain pairs. Identifies under-researched intersections where
new work could have outsized impact.

USAGE (PowerShell):
    python project1_literature_gap_analysis.py

OUTPUT:
    ../Results/literature_gap_matrix.xlsx
    ../Results/literature_gap_report.md

PROTOCOL (Pre-registered per Doctrine Section D):
    Search: PubMed via E-utilities (esearch.fcgi)
    Date range: 2020-01-01 to present
    Method: Pairwise publication count for all domain keyword pairs
    Gap score: Normalized inverse of publication count relative to
               individual domain sizes (high score = under-researched)
"""

import urllib.request
import urllib.parse
import xml.etree.ElementTree as ET
import json
import time
import os
import sys
from datetime import datetime
from itertools import combinations

# ── Configuration ──
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(SCRIPT_DIR, "..", "Results")
os.makedirs(RESULTS_DIR, exist_ok=True)

PUBMED_BASE = "https://eutils.ncbi.nlm.nih.gov/entrez/eutils/esearch.fcgi"
MIN_DATE = "2020/01/01"
MAX_DATE = datetime.now().strftime("%Y/%m/%d")
RATE_LIMIT = 0.35  # seconds between requests (NCBI allows 3/sec without API key)

# ── Research Domains & Keywords ──
# Each domain has a PubMed-optimized search term
DOMAINS = {
    "Beta Cell Regen":      '"beta cell" AND (regeneration OR replacement OR "stem cell")',
    "Insulin Resistance":   '"insulin resistance" AND (mechanism OR pathway OR molecular)',
    "Autoimmunity T1D":     '"type 1 diabetes" AND (autoimmune OR autoimmunity OR "T cell")',
    "Islet Transplant":     '"islet transplant" OR "islet cell transplant"',
    "GWAS / Polygenic":     'diabetes AND ("genome-wide" OR GWAS OR "polygenic risk")',
    "Epigenetics":          'diabetes AND (epigenetic OR methylation OR "DNA methylation")',
    "Gene Therapy":         'diabetes AND ("gene therapy" OR "gene editing" OR CRISPR)',
    "Treg / CAR-T":         'diabetes AND ("regulatory T cell" OR Treg OR "CAR-T" OR "CAR T")',
    "GLP-1 Agonists":       '"GLP-1" OR "glucagon-like peptide" AND diabetes',
    "SGLT2 Inhibitors":     '"SGLT2" AND diabetes',
    "Glucokinase":          'diabetes AND (glucokinase OR "glucose sensor")',
    "Microbiome Gut":       'diabetes AND (microbiome OR "gut microbiota" OR "gut bacteria")',
    "Personalized Nutr":    'diabetes AND ("personalized nutrition" OR "precision nutrition")',
    "Proteomics":           'diabetes AND (proteomics OR "proteomic" OR "circulating proteins")',
    "Metabolomics":         'diabetes AND (metabolomics OR metabolomic OR metabolites)',
    "Multi-Omics":          'diabetes AND ("multi-omics" OR multiomics OR "multi omics")',
    "AI / ML Predict":      'diabetes AND ("machine learning" OR "artificial intelligence" OR "deep learning")',
    "Drug Repurposing":     'diabetes AND ("drug repurposing" OR "drug repositioning")',
    "Retinopathy":          '"diabetic retinopathy"',
    "Nephropathy DKD":      '"diabetic kidney" OR "diabetic nephropathy"',
    "Neuropathy":           '"diabetic neuropathy"',
    "CV Complications":     'diabetes AND ("cardiovascular" AND complication)',
    "Closed Loop / AP":     'diabetes AND ("closed loop" OR "artificial pancreas")',
    "CGM Technology":       '"continuous glucose monitoring" OR CGM AND diabetes',
    "Health Equity":        'diabetes AND ("health equity" OR "health disparity" OR "racial disparity")',
    "Youth Diabetes":       'diabetes AND (youth OR pediatric OR adolescent) AND ("type 2")',
    "LADA":                 '"latent autoimmune diabetes" OR LADA',
    "Gestational DM":       '"gestational diabetes"',
    "Prevention / DPP":     'diabetes AND (prevention OR "prevention program")',
    "Remission T2D":        '"type 2 diabetes" AND (remission OR reversal OR "disease reversal")',
}

def query_pubmed_count(query, retries=3):
    """Query PubMed and return the count of results."""
    params = urllib.parse.urlencode({
        "db": "pubmed",
        "term": query,
        "rettype": "count",
        "datetype": "pdat",
        "mindate": MIN_DATE,
        "maxdate": MAX_DATE,
    })
    url = f"{PUBMED_BASE}?{params}"

    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": "DiabetesResearchHub/1.0"})
            with urllib.request.urlopen(req, timeout=15) as resp:
                xml_data = resp.read().decode("utf-8")
            root = ET.fromstring(xml_data)
            count_el = root.find("Count")
            if count_el is not None:
                return int(count_el.text)
            return 0
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(2)
            else:
                print(f"  [WARN] Failed after {retries} attempts: {e}")
                return -1

def run_gap_analysis():
    """Main analysis: query all domains individually, then all pairs."""
    print("=" * 60)
    print("PROJECT 1: Literature Gap Analysis")
    print(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    print(f"PubMed date range: {MIN_DATE} to {MAX_DATE}")
    print(f"Domains: {len(DOMAINS)}")
    print(f"Pairs to query: {len(DOMAINS) * (len(DOMAINS) - 1) // 2}")
    print("=" * 60)

    domain_names = list(DOMAINS.keys())

    # Step 1: Individual domain counts
    print("\n[1/3] Querying individual domain publication counts...")
    individual_counts = {}
    for i, (name, query) in enumerate(DOMAINS.items()):
        count = query_pubmed_count(query)
        individual_counts[name] = count
        print(f"  [{i+1}/{len(DOMAINS)}] {name}: {count:,} publications")
        time.sleep(RATE_LIMIT)

    # Step 2: Pairwise counts
    print("\n[2/3] Querying pairwise domain publication counts...")
    pairs = list(combinations(domain_names, 2))
    pair_counts = {}
    for i, (d1, d2) in enumerate(pairs):
        combined_query = f"({DOMAINS[d1]}) AND ({DOMAINS[d2]})"
        count = query_pubmed_count(combined_query)
        pair_counts[(d1, d2)] = count
        pair_counts[(d2, d1)] = count  # symmetric
        pct = (i + 1) / len(pairs) * 100
        if (i + 1) % 20 == 0 or i == 0 or i == len(pairs) - 1:
            print(f"  [{i+1}/{len(pairs)}] ({pct:.0f}%) {d1} x {d2}: {count}")
        time.sleep(RATE_LIMIT)

    # Step 3: Compute gap scores
    print("\n[3/3] Computing gap scores...")
    gap_scores = {}
    for d1, d2 in combinations(domain_names, 2):
        c1 = max(individual_counts.get(d1, 1), 1)
        c2 = max(individual_counts.get(d2, 1), 1)
        pair_c = max(pair_counts.get((d1, d2), 0), 0)
        # Expected overlap: geometric mean of individual proportions
        expected = (c1 * c2) ** 0.5
        # Gap score: how much less overlap exists than expected (higher = bigger gap)
        if expected > 0:
            gap_score = round(max(0, 1 - (pair_c / expected)) * 100, 1)
        else:
            gap_score = 0
        gap_scores[(d1, d2)] = {
            "domain_1": d1,
            "domain_2": d2,
            "count_d1": c1,
            "count_d2": c2,
            "pair_count": pair_c,
            "expected": round(expected, 1),
            "gap_score": gap_score,
        }

    # Sort by gap score descending
    ranked_gaps = sorted(gap_scores.values(), key=lambda x: x["gap_score"], reverse=True)

    # ── Export Results ──
    export_excel(domain_names, individual_counts, pair_counts, ranked_gaps)
    export_markdown(individual_counts, ranked_gaps)
    export_json(individual_counts, pair_counts, ranked_gaps)

    print("\n" + "=" * 60)
    print("COMPLETE. Outputs saved to Analysis/Results/")
    print("=" * 60)
    print(f"\nTop 10 Under-Researched Intersections:")
    print("-" * 55)
    for i, g in enumerate(ranked_gaps[:10], 1):
        print(f"  {i:2d}. {g['domain_1']} x {g['domain_2']}")
        print(f"      Gap Score: {g['gap_score']}  |  Publications: {g['pair_count']}  |  Expected: {g['expected']}")

def export_excel(domain_names, individual_counts, pair_counts, ranked_gaps):
    """Export gap matrix as Excel file."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("  [INFO] openpyxl not available. Skipping Excel export.")
        print("         Install with: pip install openpyxl")
        return

    wb = Workbook()
    hf = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    df = Font(name="Arial", size=9)
    border = Border(*(Side(style="thin", color="D9D9D9"),) * 4)

    # Sheet 1: Gap Matrix (heatmap style)
    ws1 = wb.active
    ws1.title = "Gap Matrix"
    n = len(domain_names)

    # Headers
    ws1.cell(row=1, column=1, value="Domain").font = hf
    ws1.cell(row=1, column=1).fill = PatternFill("solid", fgColor="1F4E79")
    ws1.column_dimensions["A"].width = 20
    for j, name in enumerate(domain_names, 2):
        c = ws1.cell(row=1, column=j, value=name)
        c.font = Font(name="Arial", bold=True, size=8, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1F4E79")
        c.alignment = Alignment(textRotation=90, horizontal="center")
        ws1.column_dimensions[get_column_letter(j)].width = 5

    # Data with conditional coloring
    for i, d1 in enumerate(domain_names, 2):
        ws1.cell(row=i, column=1, value=d1).font = Font(name="Arial", bold=True, size=9)
        for j, d2 in enumerate(domain_names, 2):
            if d1 == d2:
                val = individual_counts.get(d1, 0)
                cell = ws1.cell(row=i, column=j, value=val)
                cell.fill = PatternFill("solid", fgColor="D9E2F3")
            else:
                val = pair_counts.get((d1, d2), 0)
                cell = ws1.cell(row=i, column=j, value=val)
                # Color: fewer publications = more red (bigger gap)
                if val == 0:
                    cell.fill = PatternFill("solid", fgColor="FF0000")
                elif val < 10:
                    cell.fill = PatternFill("solid", fgColor="FF6B6B")
                elif val < 50:
                    cell.fill = PatternFill("solid", fgColor="FFA500")
                elif val < 200:
                    cell.fill = PatternFill("solid", fgColor="FFD700")
                elif val < 1000:
                    cell.fill = PatternFill("solid", fgColor="90EE90")
                else:
                    cell.fill = PatternFill("solid", fgColor="228B22")
                    cell.font = Font(name="Arial", size=8, color="FFFFFF")
            cell.font = Font(name="Arial", size=8)
            cell.alignment = Alignment(horizontal="center")
            cell.border = border

    # Sheet 2: Ranked Gaps
    ws2 = wb.create_sheet("Ranked Gaps")
    cols = ["Rank", "Domain 1", "Domain 2", "Gap Score", "Pair Publications",
            "Expected Overlap", "Domain 1 Total", "Domain 2 Total", "Opportunity"]
    widths = [6, 20, 20, 12, 16, 16, 16, 16, 14]
    for j, (col, w) in enumerate(zip(cols, widths), 1):
        c = ws2.cell(row=1, column=j, value=col)
        c.font = hf
        c.fill = PatternFill("solid", fgColor="2D6A4F")
        ws2.column_dimensions[get_column_letter(j)].width = w

    for i, g in enumerate(ranked_gaps, 2):
        rank = i - 1
        opp = "HIGH" if g["gap_score"] >= 90 else "MEDIUM" if g["gap_score"] >= 70 else "LOW"
        row = [rank, g["domain_1"], g["domain_2"], g["gap_score"],
               g["pair_count"], g["expected"], g["count_d1"], g["count_d2"], opp]
        for j, val in enumerate(row, 1):
            cell = ws2.cell(row=i, column=j, value=val)
            cell.font = df
            cell.border = border
            if j == 9:
                colors = {"HIGH": "FF6B6B", "MEDIUM": "FFA500", "LOW": "90EE90"}
                cell.fill = PatternFill("solid", fgColor=colors.get(val, "FFFFFF"))

    ws2.auto_filter.ref = f"A1:I{len(ranked_gaps)+1}"
    ws2.freeze_panes = "A2"

    # Sheet 3: Individual Domain Counts
    ws3 = wb.create_sheet("Domain Counts")
    ws3.cell(row=1, column=1, value="Domain").font = hf
    ws3.cell(row=1, column=1).fill = PatternFill("solid", fgColor="4A148C")
    ws3.cell(row=1, column=2, value="Publications (2020+)").font = hf
    ws3.cell(row=1, column=2).fill = PatternFill("solid", fgColor="4A148C")
    ws3.column_dimensions["A"].width = 22
    ws3.column_dimensions["B"].width = 20
    for i, (name, count) in enumerate(sorted(individual_counts.items(), key=lambda x: -x[1]), 2):
        ws3.cell(row=i, column=1, value=name).font = df
        ws3.cell(row=i, column=2, value=count).font = df

    out_path = os.path.join(RESULTS_DIR, "literature_gap_matrix.xlsx")
    wb.save(out_path)
    print(f"  Saved: {out_path}")

def export_markdown(individual_counts, ranked_gaps):
    """Export gap report as Markdown."""
    lines = [
        "# Literature Gap Analysis Report",
        f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"**Source:** PubMed (date range: {MIN_DATE} to {MAX_DATE})",
        f"**Domains analyzed:** {len(DOMAINS)}",
        f"**Pairs analyzed:** {len(DOMAINS) * (len(DOMAINS) - 1) // 2}",
        "",
        "---",
        "",
        "## Top 25 Under-Researched Intersections",
        "",
        "These domain pairs have significantly fewer joint publications than expected,",
        "suggesting under-explored research territory where new work could fill gaps.",
        "",
        "| Rank | Domain 1 | Domain 2 | Gap Score | Joint Pubs | Expected | Opportunity |",
        "|------|----------|----------|-----------|------------|----------|-------------|",
    ]
    for i, g in enumerate(ranked_gaps[:25], 1):
        opp = "HIGH" if g["gap_score"] >= 90 else "MEDIUM" if g["gap_score"] >= 70 else "LOW"
        lines.append(
            f"| {i} | {g['domain_1']} | {g['domain_2']} | {g['gap_score']} | "
            f"{g['pair_count']} | {g['expected']} | {opp} |"
        )

    lines += [
        "",
        "---",
        "",
        "## Interpretation Guide",
        "",
        "**Gap Score** (0-100): Measures how much less overlap exists between two domains",
        "compared to what would be expected given each domain's individual publication volume.",
        "A score of 95 means the intersection has 95% fewer publications than expected.",
        "",
        "**Opportunity Levels:**",
        "- **HIGH** (90+): Very few publications at this intersection. Likely unexplored territory.",
        "- **MEDIUM** (70-89): Some publications exist but well below expected. Room for contribution.",
        "- **LOW** (<70): Reasonably well-covered. New work here competes with existing literature.",
        "",
        "---",
        "",
        "## Individual Domain Publication Volumes",
        "",
        "| Domain | Publications (2020+) |",
        "|--------|---------------------|",
    ]
    for name, count in sorted(individual_counts.items(), key=lambda x: -x[1]):
        lines.append(f"| {name} | {count:,} |")

    lines += [
        "",
        "---",
        "",
        "## Validation Notes",
        "",
        "- Source: PubMed E-utilities API (esearch.fcgi)",
        "- All counts are approximate (PubMed search matching, not exact MeSH)",
        "- Gap scores are relative measures, not absolute judgments",
        "- Low publication count may indicate: (a) genuinely unexplored territory,",
        "  (b) terminology mismatch, or (c) research published under different keywords",
        "- This analysis should be cross-referenced with expert domain knowledge",
        "  before drawing conclusions",
        "",
        f"*Generated by Diabetes Research Hub — Project 1 — {datetime.now().strftime('%Y-%m-%d')}*",
    ]

    out_path = os.path.join(RESULTS_DIR, "literature_gap_report.md")
    with open(out_path, "w", encoding="utf-8") as f:
        f.write("\n".join(lines))
    print(f"  Saved: {out_path}")

def export_json(individual_counts, pair_counts, ranked_gaps):
    """Export raw data as JSON for downstream analysis."""
    data = {
        "metadata": {
            "generated": datetime.now().isoformat(),
            "source": "PubMed E-utilities",
            "date_range": f"{MIN_DATE} to {MAX_DATE}",
            "domains": len(DOMAINS),
        },
        "individual_counts": individual_counts,
        "pair_counts": {f"{k[0]}||{k[1]}": v for k, v in pair_counts.items()},
        "ranked_gaps": ranked_gaps,
    }
    out_path = os.path.join(RESULTS_DIR, "literature_gap_data.json")
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2)
    print(f"  Saved: {out_path}")

if __name__ == "__main__":
    run_gap_analysis()
